"""Result directory / CSV / JSON writer, following PNPWLS csv_writer.py and
json_results.py conventions, enforcing the standard DCPerf result layout::

    results/
    └── <workload>/
        └── <experiment>/
            └── session_<NNN>_<YYYYMMDD_HHMMSS>/
                ├── stdout.log
                ├── stderr.log
                ├── metrics.json
                ├── results.csv
                ├── results.json
                ├── system_metadata.json
                ├── command.txt
                └── emon/
                    ├── emon_raw/
                    ├── emon_processed/
                    └── tmc_upload.log   (only if -ue used)

``<experiment>`` defaults to ``exp_<YYYYMMDD>`` when --experiment is not
given. ``consolidated_results.xlsx`` (one sheet per workload) lives directly
under the base results dir -- see append_to_consolidated().
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import socket
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import fcntl
except ImportError:  # pragma: no cover - fcntl is POSIX-only; SUTs are Linux
    fcntl = None  # type: ignore[assignment]

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional until requirements.txt is installed
    openpyxl = None  # type: ignore[assignment]

_SCORE_LINE_RE = re.compile(r"^(\w+):\s*([\d.]+),\s*single data point", re.MULTILINE)
_OVERALL_SCORE_RE = re.compile(r"DCPerf overall score:\s*([\d.]+)", re.IGNORECASE)
_SESSION_DIR_RE = re.compile(r"^session_(\d+)_")

CONSOLIDATED_COLUMNS = [
    "session_id", "experiment", "timestamp", "host", "kernel", "cpu_model", "core_count",
    "primary_kpi", "kpi_unit", "p50_latency_ms", "p99_latency_ms", "status",
    "emon_collected", "tmc_uploaded", "tmc_link", "session_path", "notes",
]


def collect_dcperf_score(dcperf_root: Optional[str], logger) -> Dict[str, Any]:
    """Run `./benchpress_cli.py report score --all` and parse its output.

    Returns {} if dcperf_root is unset, the command fails, or not enough
    benchmarks have run for an overall score -- never raises.
    """
    if not dcperf_root:
        return {}

    cli = Path(dcperf_root) / "benchpress_cli.py"
    if not cli.exists():
        logger.warning("result_manager: %s not found, cannot collect DCPerf score", cli)
        return {}

    cmd = [sys.executable, str(cli), "report", "score", "--all"]
    logger.info("result_manager: %s", " ".join(cmd))
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    except (subprocess.SubprocessError, OSError) as exc:
        logger.warning("result_manager: report score failed to run: %s", exc)
        return {}

    output = result.stdout or ""
    scores: Dict[str, Any] = {}
    for match in _SCORE_LINE_RE.finditer(output):
        scores[match.group(1)] = float(match.group(2))

    overall_match = _OVERALL_SCORE_RE.search(output)
    if overall_match:
        scores["overall"] = float(overall_match.group(1))

    if not scores:
        logger.warning("result_manager: no DCPerf score found in report score output")
        return {}

    scores["raw_output"] = output
    return scores


class ResultManager:
    """Owns the shared top-level results timestamp for one master run."""

    def __init__(self, base_results_dir: Path, logger):
        self.base_results_dir = Path(base_results_dir)
        self.logger = logger
        # Fixed once so every workload run in this master invocation shares
        # the same top-level timestamp for grouping in the summary report.
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ------------------------------------------------------------------
    # Directory creation
    # ------------------------------------------------------------------

    def create_run_dir(self, workload: str, session: Optional[str] = None, experiment: Optional[str] = None) -> Path:
        """Create results/<workload>/<experiment>/[session]/session_<NNN>_<timestamp>/.

        ``experiment`` defaults to ``exp_<YYYYMMDD>`` (sanitized). ``session``
        is the pre-existing corescaling-sweep grouping folder, kept as an
        extra path segment under the experiment when supplied.
        """
        experiment_name = self._sanitize_name(experiment) if experiment else f"exp_{datetime.now().strftime('%Y%m%d')}"
        parent_dir = self.base_results_dir / workload / experiment_name
        if session:
            parent_dir = parent_dir / session
        run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = self._build_session_id(parent_dir, run_timestamp)
        run_dir = parent_dir / session_id
        suffix = 1
        while run_dir.exists():
            run_dir = parent_dir / f"{session_id}_{suffix}"
            suffix += 1
        try:
            run_dir.mkdir(parents=True, exist_ok=True)
            (run_dir / "emon").mkdir(parents=True, exist_ok=True)
            (run_dir / "emon" / "emon_raw").mkdir(parents=True, exist_ok=True)
            (run_dir / "emon" / "emon_processed").mkdir(parents=True, exist_ok=True)
        finally:
            # Printed on both success and failure paths (Gate C contract).
            print(f"Output Directory: {run_dir.resolve()}")
        return run_dir

    def _build_session_id(self, parent_dir: Path, timestamp: str) -> str:
        """Count existing session_* dirs in parent_dir; return session_<NNN+1>_<timestamp>."""
        existing = 0
        if parent_dir.exists():
            for child in parent_dir.iterdir():
                if child.is_dir() and _SESSION_DIR_RE.match(child.name):
                    existing += 1
        return f"session_{existing + 1:03d}_{timestamp}"

    @staticmethod
    def _sanitize_name(name: str) -> str:
        """Lowercase, replace non-alphanumeric chars with underscore, truncate to 64 chars."""
        sanitized = re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_")
        return (sanitized or "exp")[:64]

    def get_emon_raw_dir(self, run_dir: Path) -> Path:
        """Create (if needed) and return <run_dir>/emon/emon_raw/."""
        path = Path(run_dir) / "emon" / "emon_raw"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def get_emon_processed_dir(self, run_dir: Path) -> Path:
        """Create (if needed) and return <run_dir>/emon/emon_processed/."""
        path = Path(run_dir) / "emon" / "emon_processed"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def write_tmc_upload_log(self, run_dir: Path, content: str, tmc_link: str = "") -> Path:
        """Write emon/tmc_upload.log with upload status and optional TMC URL."""
        log_path = Path(run_dir) / "emon" / "tmc_upload.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        text = content or ""
        if tmc_link:
            text += f"\ntmc_link: {tmc_link}\n"
        log_path.write_text(text, encoding="utf-8")
        self.logger.debug("result_manager: wrote %s", log_path)
        return log_path

    # ------------------------------------------------------------------
    # Simple artifact writers
    # ------------------------------------------------------------------

    def save_stdout(self, run_dir: Path, content: str) -> None:
        (Path(run_dir) / "stdout.log").write_text(content or "", encoding="utf-8")

    def save_stderr(self, run_dir: Path, content: str) -> None:
        (Path(run_dir) / "stderr.log").write_text(content or "", encoding="utf-8")

    def save_command(self, run_dir: Path, cmd: str) -> None:
        (Path(run_dir) / "command.txt").write_text(cmd or "", encoding="utf-8")

    def save_metrics(self, run_dir: Path, metrics: Dict[str, Any]) -> None:
        (Path(run_dir) / "metrics.json").write_text(
            json.dumps(metrics, indent=2, default=str), encoding="utf-8"
        )

    # ------------------------------------------------------------------
    # CSV writer (PNPWLS csv_writer.py style: smart header handling)
    # ------------------------------------------------------------------

    def write_csv_row(self, run_dir: Path, row: Dict[str, Any]) -> None:
        """Append one row to results.csv, writing the header only if new.

        Never silently swallows failures — exceptions propagate to the
        caller so a disk-full/permission error is visible.
        """
        csv_file = Path(run_dir) / "results.csv"
        header = list(row.keys())
        values = [str(v) if v is not None else "" for v in row.values()]

        file_exists = csv_file.exists() and csv_file.stat().st_size > 0
        write_header = True
        if file_exists:
            with open(csv_file, "r", newline="", encoding="utf-8") as fh:
                first_line = fh.readline().strip()
            write_header = first_line != ",".join(header)

        with open(csv_file, "a", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
            if write_header:
                writer.writerow(header)
            writer.writerow(values)

        self.logger.info("result_manager: wrote row to %s", csv_file)

    # ------------------------------------------------------------------
    # Consolidated cross-run workbook (one sheet per workload)
    # ------------------------------------------------------------------

    def append_to_consolidated(self, workload: str, row_data: Dict[str, Any]) -> None:
        """Append one row to consolidated_results.xlsx on the <workload> sheet.

        Creates the file and/or sheet if missing. Never overwrites or deletes
        existing rows. Uses fcntl.flock on a sidecar .lock file so concurrent
        runs across workloads/hosts don't corrupt the workbook.
        """
        if openpyxl is None:
            self.logger.warning("result_manager: openpyxl not installed, skipping consolidated_results.xlsx")
            return

        xlsx_path = self.base_results_dir / "consolidated_results.xlsx"
        lock_path = self.base_results_dir / "consolidated_results.xlsx.lock"
        self.base_results_dir.mkdir(parents=True, exist_ok=True)
        row = [row_data.get(col, "") for col in CONSOLIDATED_COLUMNS]

        lock_file = open(lock_path, "w")
        try:
            if fcntl is not None:
                fcntl.flock(lock_file, fcntl.LOCK_EX)
            try:
                if xlsx_path.exists():
                    workbook = openpyxl.load_workbook(xlsx_path)
                else:
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)  # drop the default blank sheet

                if workload in workbook.sheetnames:
                    sheet = workbook[workload]
                else:
                    sheet = workbook.create_sheet(title=workload)
                    sheet.append(CONSOLIDATED_COLUMNS)

                sheet.append(row)
                workbook.save(xlsx_path)
                self.logger.debug("result_manager: appended row to %s!%s", xlsx_path, workload)
            finally:
                if fcntl is not None:
                    fcntl.flock(lock_file, fcntl.LOCK_UN)
        except OSError as exc:
            self.logger.error("result_manager: failed to update consolidated_results.xlsx: %s", exc)
        finally:
            lock_file.close()

    # ------------------------------------------------------------------
    # JSON results writer (PNPWLS json_results.py style)
    # ------------------------------------------------------------------

    def write_json_results(self, run_dir: Path, data: Dict[str, Any]) -> None:
        """Always write results.json, on both success and failure paths."""
        results_file = Path(run_dir) / "results.json"
        payload = {
            "version": 2,
            "orch_run_id": data.get("orch_run_id", ""),
            "rows": data.get("rows", []),
        }
        try:
            results_file.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        except OSError as exc:
            self.logger.error("result_manager: FAILED to write results.json: %s", exc)
            raise

    def write_system_metadata(self, run_dir: Path, metadata: Dict[str, Any]) -> None:
        (Path(run_dir) / "system_metadata.json").write_text(
            json.dumps(metadata, indent=2, default=str), encoding="utf-8"
        )

    # ------------------------------------------------------------------
    # benchpress benchmark_metrics_<run_id>/ preservation
    # ------------------------------------------------------------------

    def copy_benchmark_metrics(self, dcperf_root: str, run_id: str, run_dir: Path) -> bool:
        """Copy benchpress's own `benchmark_metrics_<run_id>/` folder into run_dir.

        benchpress writes per-instance CSVs/logs (feedsim per-instance QPS,
        tao_bench per-server CSVs, video per-level results, etc.) into this
        folder under dcperf_root -- previously left behind and never
        captured into our organized results/ directory.
        """
        src = Path(dcperf_root) / f"benchmark_metrics_{run_id}"
        if not src.exists():
            self.logger.warning("result_manager: %s not found, nothing to preserve", src)
            return False

        dest = Path(run_dir) / "benchmark_metrics"
        try:
            shutil.copytree(src, dest, dirs_exist_ok=True)
            self.logger.info("result_manager: copied %s -> %s", src, dest)
            return True
        except OSError as exc:
            self.logger.error("result_manager: failed to copy %s: %s", src, exc)
            return False

    # ------------------------------------------------------------------
    # Master run summary
    # ------------------------------------------------------------------

    def write_summary(self, all_results: List[Dict[str, Any]], dcperf_score: Optional[Dict[str, Any]] = None) -> None:
        """Write run_summary.json and run_summary.txt into the top-level results dir."""
        summary_dir = self.base_results_dir / f"summary_{self.timestamp}"
        summary_dir.mkdir(parents=True, exist_ok=True)

        (summary_dir / "run_summary.json").write_text(
            json.dumps({"results": all_results, "dcperf_score": dcperf_score or {}}, indent=2, default=str),
            encoding="utf-8",
        )

        text = self._render_summary_text(all_results, dcperf_score)
        (summary_dir / "run_summary.txt").write_text(text, encoding="utf-8")
        print(text)

    def _render_summary_text(self, all_results: List[Dict[str, Any]], dcperf_score: Optional[Dict[str, Any]] = None) -> str:
        hostname = socket.gethostname()
        cpu_model = self._read_cpu_model()
        bar = "=" * 64
        thin = "-" * 64

        lines = [
            bar,
            "DCPerf Run Summary",
            f"Timestamp : {self.timestamp}",
            f"Host      : {hostname}",
            f"Platform  : {cpu_model}",
            bar,
            f"{'Workload':<22}{'Status':<8}{'Runs':<7}{'Primary KPI'}",
            thin,
        ]

        pass_count = 0
        fail_count = 0
        for entry in all_results:
            status = entry.get("status", "FAIL")
            if status == "PASS":
                pass_count += 1
            else:
                fail_count += 1
            kpi = entry.get("primary_kpi", "--")
            lines.append(
                f"{entry.get('workload', ''):<22}{status:<8}{str(entry.get('runs', '')):<7}{kpi}"
            )

        lines.append(thin)
        total = len(all_results)
        lines.append(f"Total: {total} workloads | {pass_count} PASS | {fail_count} FAIL")
        lines.append(f"Results: results/{self.timestamp}/")
        lines.append(bar)

        lines.append("")
        lines.append(bar)
        lines.append("DCPerf Benchmark Scores")
        lines.append(bar)
        per_benchmark = {k: v for k, v in (dcperf_score or {}).items() if k not in ("overall", "raw_output")}
        if per_benchmark:
            for name, value in per_benchmark.items():
                lines.append(f"{name:<14}: {value}")
            lines.append(thin)
            overall = (dcperf_score or {}).get("overall")
            if overall is not None:
                lines.append(f"DCPerf Overall Score : {overall}")
        else:
            lines.append("DCPerf Score: Not available (run all 5 benchmarks for overall score)")
        lines.append(bar)
        return "\n".join(lines) + "\n"

    @staticmethod
    def _read_cpu_model() -> str:
        try:
            with open("/proc/cpuinfo", encoding="utf-8", errors="ignore") as fh:
                for line in fh:
                    if line.lower().startswith("model name"):
                        return line.split(":", 1)[1].strip()
        except OSError:
            pass
        return "Unknown"
